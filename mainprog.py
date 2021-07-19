import openpyxl as xl
import xlsxwriter
from datetime import datetime


# class for car rental
class CarRental:
    _car_type = None
    _total_hour = None
    _cust_id = None
    _date = None
    _st_time = None
    _cust_name = None
    _rental_mode = None
    _car_number = None

    def __init__(self):
        # print('Construstor Car Rental')
        pass

    def create_CarBooking_xls(self):

        # Create a workbook and add customer details
        workbook = xlsxwriter.Workbook('CarBooking.xlsx')
        worksheet = workbook.add_worksheet()

        # Some data we want to write to the worksheet.
        car_rental_details = (
            ['car_number', 'car_type', 'date', 'st_time', 'total_hours', 'cust_ID', 'Rental_Mode'],
            ['DL7C4200', 'Sedan', '27/11/01', '10 am', '5', 'C001', 'Hourly'],
            ['DL7B0007', 'Luxury', '27/11/01', '11 am', '2', 'C002', 'Weekly'],
            ['DL5CC9922', 'Sedan', '27/11/01', '03 pm', '8', 'C003', 'Monthly']
        )

        # Start from the first cell. Rows and columns are zero indexed.
        row = 0

        # Iterate over the data and write it out row by row.
        for car_numb, car_type, date, st_time, hours, cust_ID, Rental_Mode in car_rental_details:
            worksheet.write(row, 0, car_numb)
            worksheet.write(row, 1, car_type)
            worksheet.write(row, 2, date)
            worksheet.write(row, 3, st_time)
            worksheet.write(row, 4, hours)
            worksheet.write(row, 5, cust_ID)
            worksheet.write(row, 6, Rental_Mode)
            row += 1

        workbook.close()

    def car_display(self):
        wb = xl.load_workbook('CarRentalDetails.xlsx')
        sheet = wb['Sheet1']
        for row in sheet.iter_rows(values_only=True):
            print(row)

    def car_booking(self):
        print("\n________________________________")
        print("Welcome to the Car rental portal")
        print("________________________________\n")
        isvalid = False
        while not isvalid:
            print("Choose options from the following :\n")
            print("1. Show available cars ")
            print("2. Book the car for rental ")
            print("3. Check are availablity basis car type ")
            print("4. Exit \n")
            num = int(input("Enter your choice "))
            try:
                if num == 1:
                    self.car_display()
                elif num == 2:
                    customer().accept_cust_request()
                elif num == 3:
                    print("step 3")
                    c_type = input("Enter the car category to be checked (Sedan/SUV/Luxury/Hatchback/Max) ")
                    Car().car_validate(c_type)
                elif num == 4:
                    isvalid = True
                else:
                    print("Please choose the correct option, try again!\n")
            except:
                print("\n")

    def rental_car_monthly(self, CarRental):
        booking_date, booking_time = customer().obtain_date()
        _total_hour = input("Enter number of months : ")
        _car_number = input("Enter car number : ")
        wb = xl.load_workbook('CarBooking.xlsx')
        sheet = wb.worksheets[0]
        _rental_mode = 'Monthly'
        data = ([_car_number, self._car_type, booking_date, booking_time, _total_hour, self._cust_id, _rental_mode])
        sheet.append(data)
        wb.save('CarBooking.xlsx')
        print("monthly executed")
        wb.close()

    def rental_car_weekly(self, CarRental):
        booking_date, booking_time = customer().obtain_date()
        _total_hour = input("Enter number of weeks : ")
        _car_number = input("Enter car number : ")
        wb = xl.load_workbook('CarBooking.xlsx')
        sheet = wb.worksheets[0]
        _rental_mode = 'Weekly'
        data = ([_car_number, self._car_type, booking_date, booking_time, _total_hour, self._cust_id, _rental_mode])
        sheet.append(data)
        wb.save('CarBooking.xlsx')
        print("Weekly executed")
        wb.close()

    def rental_car_hourly(self, CarRental):
        booking_date, booking_time = customer().obtain_date()
        _total_hour = input("Enter number of hours : ")
        _car_number = input("Enter car number : ")
        wb = xl.load_workbook('CarBooking.xlsx')
        sheet = wb.worksheets[0]
        _rental_mode = 'Hourly'
        data = ([_car_number, self._car_type, booking_date, booking_time, _total_hour, self._cust_id, _rental_mode])
        sheet.append(data)
        wb.save('CarBooking.xlsx')
        print("Hourly executed")
        wb.close()

class customer:
    _date = None

    def __init__(self):
        pass

    def add_customer(self):
        pass

    def create_customer_xls(self):
        # Create a workbook and add customer details
        workbook = xlsxwriter.Workbook('Customer.xlsx')
        worksheet = workbook.add_worksheet()

        # Some data we want to write to the worksheet.
        cust_details = (
            ['cust_id', 'cust_name', 'cust_phone', 'cust_address', 'cust_email'],
            ['C001', 'Neha', '2020699887', 'Delhi', 'Neha@abc.com'],
            ['C002', 'Akshey', '2019878987', 'Delhi', 'Ak@xyz.com'],
            ['C003', 'Yashmit', '9056432019', 'Okhla', 'Yash@abc.com']
        )

        # Start from the first cell. Rows and columns are zero indexed.
        row = 0

        # Iterate over the data and write it out row by row.
        for c_id, c_name, c_phone, address, email in cust_details:
            worksheet.write(row, 0, c_id)
            worksheet.write(row, 1, c_name)
            worksheet.write(row, 2, c_phone)
            worksheet.write(row, 3, address)
            worksheet.write(row, 4, email)
            row += 1
        workbook.close()

    def accept_cust_details(self):
        c_id = input("Customer ID ")
        if self.validate_customer(c_id):
            print("Customer already exist")
            c_id = input("Enter new customer ID ")
        c_name = input("Customer Name ")
        c_phone = input("Customer Phone No. ")
        address = input("Customer Address ")
        email = input("Customer email ")

    def cust_add(self, c_id, c_name, c_phone, address, email):
        wb = xl.load_workbook('Customer.xlsx')
        sheet = wb.worksheets[0]
        data = ([c_id, c_name, c_phone, address, email])
        sheet.append(data)
        wb.save('Customer.xlsx')
        wb.close()

    def disp_cust_details(self):
        wb = xl.load_workbook('Customer.xlsx')
        sheet = wb['Sheet1']
        for row in sheet.iter_rows(values_only=True):
              print(row)

    def disp_customer_detail(self, cust_id):
        wb = xl.load_workbook('Customer.xlsx')
        sheet = wb['Sheet1']
        for row in range(1, sheet.max_row + 1):
            for column in "A":
                if sheet.cell(row, column=1).value == cust_id:
                    for x in range(1, sheet.max_column + 1):
                       cell = sheet.cell(row, x)
                       print(cell.value, end='\t\t')
                    break
                break

    def validate_customer(self, cust_id):
        val = False
        wb = xl.load_workbook('Customer.xlsx')
        sheet = wb['Sheet1']
        for row in range(1, sheet.max_row + 1):
            for column in "A":
                if sheet.cell(row, column=1).value == cust_id:
                    val = True
                    break
                break
        return val

    # Function to accept date from the user
    def obtain_date(self):
        isvalid = False
        while not isvalid:
            my_string = str(input("Enter date(yyyy-mm-dd hh:mm): "))
            try:
                _date = datetime.strptime(my_string, "%Y-%m-%d %H:%M")
                booking_time = str(_date.strftime("%H:%M:%S"))
                booking_date = str(_date.strftime("%m/%d/%Y"))
                isvalid = True
            except:
                print("Please enter the date and time as per the format given, try again!\n")
        return booking_date, booking_time

    # Function to start the Project
    def accept_cust_request(self):
        print('\n_______________________________________')
        print('Welcome to the Car rental Application :')
        print('_______________________________________\n')
        new_rental = CarRental()
        new_rental._cust_id = input("Enter Customer ID : ")
        if customer().validate_customer(new_rental._cust_id) == False:
            print("Customer does not exist")
            print(("Enter Customer details for the new Customer "))
            print("_____________________________________________\n")
        new_rental._car_type = input("Enter car type to be rented (Sedan/SUV/Luxury/Hatchback/Max) : ")
        new_rental._rental_mode = input("Enter the mode of booking (Hourly/Weekly/Monthly : ")
        if new_rental._rental_mode == "Hourly" or new_rental._rental_mode == "hourly":
            print("Hourly if executed")
            new_rental.rental_car_hourly(new_rental)
            print('Record Inserted')
        elif new_rental._rental_mode == "Weekly" or new_rental._rental_mode == "weekly":
            print("Weekly if executed")
            new_rental.rental_car_weekly(new_rental)
            print('Record Inserted')
        elif new_rental._rental_mode == "Monthly" or new_rental._rental_mode == "monthly":
            print("Monthly if executed")
            new_rental.rental_car_monthly(new_rental)
            print('Record Inserted')
        else:
            print("Choose between Hourly/Weekly/Monthly only")



class Car:
    def __init__(self):
        print("car class")
    def create_sheet(self):
        # Create a workbook and add a worksheet.
        workbook = xlsxwriter.Workbook('CarRentalDetails.xlsx')
        worksheet = workbook.add_worksheet()

        # Some data we want to write to the worksheet.
        car_details = (
            ['car_number', 'car_type', 'Rented', 'Rental_Mode'],
            ['DL7C4200', 'Sedan', 'Yes', 'Hourly'],
            ['DL7B7689', 'Hatchback', 'No', 'None'],
            ['DL7B0007', 'Luxury', 'Yes', 'Weekly'],
            ['DL7C4378', 'Max', 'No', 'None'],
            ['DL7D3422', 'SUV', 'No', 'None']
        )

        # Start from the first cell. Rows and columns are zero indexed.
        row = 0

        # Iterate over the data and write it out row by row.
        for c_no, c_type, rented, r_mode in car_details:
            worksheet.write(row, 0, c_no)
            worksheet.write(row, 1, c_type)
            worksheet.write(row, 2, rented)
            worksheet.write(row, 3, r_mode)
            row += 1
        workbook.close()

    def car_disp(self):
        wb = xl.load_workbook('CarRentalDetails.xlsx')
        sheet = wb.worksheets[0]
        row = sheet.max_row + 1
        print(row)
        for row in sheet.values:
            for value in row:
                print(value, end='\t\t')
            print(' ')
        wb.close()

    def car_add(self, car_num, car_typ, rented='No', rental_mode='None'):
        wb = xl.load_workbook('CarRentalDetails.xlsx')
        sheet = wb.worksheets[0]
        data = ([car_num, car_typ, rented, rental_mode])
        sheet.append(data)
        wb.save('CarRentalDetails.xlsx')
        wb.close()

    def car_validate(self, car_type):
        wb = xl.load_workbook('CarRentalDetails.xlsx', read_only=True)
        sheet = wb.worksheets[0]
        row_to_show = sheet.max_row + 1
        for row in range(1, sheet.max_row + 1):
            for column in "C":  # Here you can add or reduce the columns
                if (sheet.cell(row, column=2).value == car_type):
                     cell_name = "{}{}".format(column, row)
                     if sheet[cell_name].value == 'No':
                         for x in range(1, sheet.max_column + 1):
                             cell = sheet.cell(row, x)
                             print(cell.value, end='\t\t')
                         print('')
        wb.close()



'''
create xls files
CarRental().create_CarBooking_xls()
Car().create_sheet()
customer().create_customer_xls()
'''

# function to check car availabilty
# car.car_validate('Sedan')

# Execute the code
# Ncar = CarRental()
# Ncar.car_booking()

customer().accept_cust_details()
